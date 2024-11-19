import openpyxl

path = "Book1.xlsx"

wb_obj = openpyxl.load_workbook(path)
# print(wb_obj)
sheet_obj = wb_obj.active
# print(sheet_obj)

row = sheet_obj.max_row
column = sheet_obj.max_column

# print("total rows:" , row)
# print("total columns:" , column)


i=1
while i<=row:
    j=1
    while j<=column:
        print("row:",i , "col:",j)
        cell_obj = sheet_obj.cell(row=i, column=j)
        print(cell_obj.value ,end="")
        j=j+1
        print(end='\n')
    i=i+1    